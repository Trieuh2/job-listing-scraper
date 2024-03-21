import re
import datetime
import json
import os
import hashlib
from functools import reduce
from typing import List, Dict, Union, cast

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.styles import Font, PatternFill, NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

def build_indeed_url(position: str, location: str, experience_level: str, job_type: str, max_days_posted_ago: str) -> str:
    """
    Builds a URL for Indeed job search based on the given parameters.

    Args:
        position (str): The job position to search for.
        location (str): The location where the job is based.
        experience_level (str): The experience level required for the job.
        job_type (str): The type of job (e.g., full-time, part-time).
        max_days_posted_ago (str): The maximum number of days since the job was posted.

    Returns:
        str: The constructed URL for the Indeed job search.
    """
    base_url = 'https://www.indeed.com/jobs?'
    params = []

    if position:
        params.append(f"q={position}")
    if location:
        params.append(f"l={location}")
    if experience_level and job_type:
        params.append(f"sc=0kf%3Aexplvl{experience_level}jt{job_type}%3B")
    elif experience_level:
        params.append(f"sc=0kf%3Aexplvl{experience_level}%3B")
    elif job_type:
        params.append(f"sc=0kf%3Ajt{job_type}%3B")
    if max_days_posted_ago:
        params.append(f"fromage={max_days_posted_ago}")

    return base_url + '&'.join(params)

def exclude_based_on_title(excluded_keywords: List[str], title_element_text: str) -> bool:
    """
    Returns True if the job title contains any excluded keywords.

    Args:
        excluded_keywords (List[str]): A list of keywords to be excluded from the title.
        title_element_text (str): The title of the job extracted from the HTML element.

    Returns:
        bool: True if the job title contains any of the excluded keywords, False otherwise.
    """
    cleaned_title = re.sub(r'[^a-zA-Z0-9]', ' ', title_element_text)
    title_words = cleaned_title.split()
    for idx, word in enumerate(title_words):
        if word.lower() in excluded_keywords:
            return True
        if idx < len(title_words) - 1:
            combined_word = f"{word} {title_words[idx + 1]}".lower()
            if combined_word in excluded_keywords:
                return True
    return False

def get_next_page_url(url: str) -> str:
    """
    Returns the URL for the next page of job listings.

    Args:
        url (str): The URL of the current page of job listings.

    Returns:
        str: The URL for the next page of job listings.
    """
    start_tag = "&start="
    start_index = url.find(start_tag)
    if start_index == -1:
        return url + f"{start_tag}10"
    else:
        end_index = start_index + len(start_tag)
        while end_index < len(url) and url[end_index].isdigit():
            end_index += 1
        current_page = int(url[start_index + len(start_tag):end_index])
        return url[:start_index] + f"{start_tag}{current_page + 10}" + url[end_index:]

def read_jobs_excel(filename: str) -> Dict[str, Dict[str, Union[str, int, float]]]:
    """
    Reads job records from an Excel file and returns a dictionary of data.

    Args:
        filename (str): The name of the Excel file containing job records.

    Returns:
        Dict[str, Dict[str, Union[str, int, float]]]: A dictionary where each key is a hash_id and the value is a
        dictionary of job record fields.
    """
    if not os.path.isfile(filename):
        return {}

    with open('config.json') as config_file:
        config = json.load(config_file)

    data = {}  # hash_id : record
    wb = load_workbook(filename)
    ws = cast(Worksheet, wb.active)
    config_headers = config['csv_settings']['csv_headers']
    current_headers_in_file = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(current_headers_in_file, row))
        formatted_record = {}

        for header in config_headers:
            formatted_record[header] = record.get(header, '')

        data[formatted_record['hash_id']] = formatted_record
    return data

def write_jobs_excel(filename: str, job_records: Dict[str, Dict]) -> None:
    """
    Writes job records to an Excel file.

    Args:
        filename (str): The name of the Excel file where job records will be written.
        job_records (Dict[str, Dict]): A dictionary of job records, where each key is a hash_id and the value is a
        dictionary containing the job record fields.
    """
    print("Updating Excel record data")

    with open('config.json') as config_file:
        config = json.load(config_file)

    # Check if the file exists
    file_exists = os.path.isfile(filename)
    fieldnames = config['csv_settings']['csv_headers']

    # Load the existing Worksheet or create a new Worksheet
    if file_exists:
        wb = load_workbook(filename)
        worksheet = cast(Worksheet, wb.active)
        clear_all_cell_values(worksheet)
        # Set headers
        for col, header in enumerate(fieldnames, start=1):
            worksheet.cell(row=1, column=col).value = header
    else:
        wb = Workbook()
        worksheet = cast(Worksheet, wb.active)
        worksheet.append(fieldnames)

    write_new_cell_data(worksheet, fieldnames, job_records)
    apply_worksheet_conditional_formatting(worksheet)
    update_or_create_worksheet_table(worksheet, fieldnames)
    wb.save(filename)
    print("Done updating Excel records")

def write_new_cell_data(worksheet: Worksheet, fieldnames: List[str], job_records: Dict[str, Dict]) -> None:
    """
    Writes the sorted job record data to the Worksheet.

    Args:
        worksheet (Worksheet): The Worksheet object where job records will be written.
        fieldnames (List[str]): A list of field names that correspond to the columns in the Worksheet.
        job_records (Dict[str, Dict]): A dictionary of job records, where each key is a hash_id and the value is a
        dictionary containing the job record fields.
    """
    # Sort the job records first by 'posted_date' from newest to oldest, then by 'company' in alphabetical order
    sorted_job_records = sorted(job_records.values(), key=lambda x: x.get('company', ''))
    sorted_job_records.sort(key=lambda x: x['posted_date'], reverse=True)

    # Write the new data starting from row 2
    row_num = 2
    for job_record in sorted_job_records:
        col_num = 1
        for header in fieldnames:
            cell = worksheet.cell(row=row_num, column=col_num, value=job_record.get(header, ''))

            if header == 'job_link' and cell.value:  # Check if the column is 'job_link' and has a value
                setattr(cell, "hyperlink", cell.value) # Set the hyperlink
                cell.style = 'Hyperlink'  # Apply the hyperlink style
            col_num += 1
        row_num += 1

def clear_all_cell_values(worksheet: Worksheet) -> None:
    """
    Clears all cell values in the worksheet.

    Args:
        worksheet (Worksheet): The Worksheet object from which all cell values will be cleared.
    """
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

            if cell.style != 'Normal':
                cell.style = 'Normal'

def update_or_create_worksheet_table(worksheet: Worksheet, fieldnames: List[str]) -> None:
    """
    Creates or updates the table reference for the job data in the worksheet.

    Args:
        worksheet (Worksheet): The Worksheet object where the job data table will be created or updated.
        fieldnames (List[str]): A list of field names that correspond to the columns in the Worksheet.
    """
    table_exists = len(worksheet.tables) > 0
    if table_exists:
        table = worksheet.tables[next(iter(worksheet.tables))]
        table.ref = f"A1:{chr(64 + len(fieldnames))}{worksheet.max_row}"
        table.tableStyleInfo.showRowStripes = True 
    else:
        table = Table(displayName="JobTable", ref=f"A1:{chr(64 + len(fieldnames))}{worksheet.max_row}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        worksheet.add_table(table)

def apply_worksheet_conditional_formatting(worksheet: Worksheet) -> None:
    """
    Applies conditional formatting to column B, used for the 'applied' column in the worksheet.

    Cells marked as "Yes" will be filled green.
    Cells marked as "No" will be marked red.
    Cells marked as "Skip" will be marked yellow.

    Args:
        worksheet (Worksheet): The Worksheet object to which conditional formatting will be applied.
    """
    # Remove existing conditional formatting rules for the worksheet
    worksheet.conditional_formatting = ConditionalFormattingList()

    green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
    red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    worksheet.conditional_formatting.add('B2:B{}'.format(worksheet.max_row), CellIsRule(operator='equal', formula=['"Yes"'], fill=green_fill))
    worksheet.conditional_formatting.add('B2:B{}'.format(worksheet.max_row), CellIsRule(operator='equal', formula=['"No"'], fill=red_fill))
    worksheet.conditional_formatting.add('B2:B{}'.format(worksheet.max_row), CellIsRule(operator='equal', formula=['"Skip"'], fill=yellow_fill))

def string_to_hash(input_string: str) -> str:
    """
    Converts a string to a SHA-256 hash.

    Args:
        input_string (str): The string to be converted to a SHA-256 hash.

    Returns:
        str: The SHA-256 hash of the input string.
    """
    return hashlib.sha256(input_string.encode()).hexdigest()

def parse_indeed_url(url: str) -> str:
    """Parses an Indeed URL and returns the base URL."""
    second_equal_index = url.find('=', url.find('=') + 1)
    return url if second_equal_index == -1 else url[:second_equal_index]

def parse_post_date(post_date_string: str) -> str:
    """
    Parses an Indeed URL and returns the base URL.

    Args:
        url (str): The Indeed URL to be parsed.

    Returns:
        str: The base URL of the Indeed URL.
    """
    if post_date_string in ["Today", "Just posted"]:
        return datetime.date.today().strftime("%m/%d/%Y")
    elif " day ago" in post_date_string:
        return (datetime.date.today() - datetime.timedelta(days=1)).strftime("%m/%d/%Y")
    elif " days ago" in post_date_string:
        days_ago = int(post_date_string.split()[2])
        return (datetime.date.today() - datetime.timedelta(days=days_ago)).strftime("%m/%d/%Y")
    else:
        return datetime.date.today().strftime("%m/%d/%Y")

def is_valid_indeed_job_link_structure(url: str) -> bool:
    """
    Checks if an Indeed job link is valid, based on the structure of the URL.

    Args:
        url (str): The Indeed job link to be checked.

    Returns:
        bool: True if the link is a valid Indeed job link, False otherwise.
    """
    return url.startswith('https://www.indeed.com/rc/clk?jk=')

def has_valid_years_of_experience(description: str) -> bool:
    """
    Checks if the user's specified maximum years of experience meets the minimum years of experience mentioned in the job description.

    Args:
        description (str): The job description to be checked.

    Returns:
        bool: True if the job description meets the years of experience criteria, False otherwise.
    """
    # Regular expression to match variations of years of experience
    regex = r'(\d+)\+?[\s\w]* years'
    matches = re.findall(regex, description, re.IGNORECASE)

    if matches:
        years = list(map(int, matches))
        min_exp = min(years)

        with open('config.json') as config_file:
            config = json.load(config_file)

        if config['indeed_criteria']['user_years_of_experience']:
            user_years_of_experience = int(config['indeed_criteria']['user_years_of_experience'])
            if user_years_of_experience < min_exp:
                return False
        else:
            return False
    return True
    
def update_config_field(filepath: str, field_path: str, new_value: Union[str, List[str], int, bool]) -> None:
    """
    Updates a specific field in the config.json configuration file.

    Args:
        filepath (str): The path to the config.json file.
        field_path (str): The path to the field within the config that needs to be updated.
        new_value (Union[str, List[str], int, bool]): The new value to be set for the specified field.
    """
    if field_path == 'excluded_keywords' and isinstance(new_value, list):
        new_value = sorted([value.lower() for value in new_value if value.strip()])

    with open(filepath, 'r') as file:
        config = json.load(file)

    keys = field_path.split('.')
    reduce(lambda d, k: d.setdefault(k, {}), keys[:-1], config)[keys[-1]] = new_value
    with open(filepath, 'w') as file:
        json.dump(config, file, indent=4)

def is_valid_numerical_field_input(input: str) -> bool:
    """
    Checks if an input string is a valid numerical field.

    Args:
        input_str (str): The input string to be checked.

    Returns:
        bool: True if the input string is a valid numerical field, False otherwise.
    """
    return input.isdigit() or input== ""